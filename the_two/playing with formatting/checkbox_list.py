import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd

root = tk.Tk()
root.title("Checkboxes")
root.geometry('600x400')

# def clicked():
#     if var1.get() ==1:
#         my_label.config(text="Pepperoni")
#     else:
#         my_label.config(text="pick a topping")

# var1 = IntVar()
# var2 = IntVar()
# var3 = IntVar()

# check1 = Checkbutton(root, text='Pepperoni', variable=var1, onvalue=1, offvalue=0)
# check1.pack(pady=(40, 10))

# check2 = Checkbutton(root, text='Cheese', variable=var2, onvalue=1, offvalue=0)
# check2.pack(pady=(40, 10))

# check3 = Checkbutton(root, text='Mushroom', variable=var3, onvalue=1, offvalue=0)
# check3.pack(pady=(40, 10))

# my_button = Button(root, text="Submit", command=clicked)
# my_button.pack(pady=20)

# my_label = Label(root, text="pick a topping")
# my_label.pack(pady=20)


# wb = load_workbook("C:/Users/enfxm/Desktop/Python Testing/Test Report Template/Test-Report-Generator/the_two/TEMPLATE_FORMAT/cover_page/equipment_used/xls/data.xlsx")
# ws = wb.active

# col_a = ws["A"]
# col_b = ws["B"]
# col_c = ws["C"]

# my_listbox = Listbox(root, width=45)
# my_listbox.pack(pady=20)

# for item in col_a:
#     my_listbox.insert(END, item.value)



# def insert_check_btn(_range):

#     for r in range(_range):

#         var = tk.IntVar()

#         check_btn_vars.append(var)
#         check_btn = tk.Checkbutton(text_area, text=f'Item - {r}',
#                                    variable=var)

#         text_area.window_create('end', window=check_btn)
#         text_area.insert('end', '\n')

#     text_area.configure(state=tk.DISABLED)


fm = tk.Frame(root)

scroll_bar = tk.Scrollbar(fm)
scroll_bar.pack(side=tk.RIGHT, fill=tk.Y)

text_area = tk.Text(fm, bg='SystemButtonFace')
text_area.pack()

text_area.configure(yscrollcommand=scroll_bar.set)
scroll_bar.configure(command=text_area.yview)

fm.pack()



# insert_check_btn(50)

path = r"C:\Users\enfxm\Desktop\Python Testing\Test Report Template\Test-Report-Generator\the_two\TEMPLATE FORMAT\cover_page\equipment_used\csv\data.csv"
check_btn_vars = []


# make a checklist button for each row in the csv file, the name of the button should be col1 - col2 - col3 - col4 - col5 - col6 - col7
df = pd.read_csv(path)
for index, row in df.iterrows():
    var = tk.IntVar()
    check_btn_vars.append(var)
    
    # Create a string for the button text
    button_text = ' - '.join(str(row[col]) for col in df.columns)
    
    check_btn = tk.Checkbutton(text_area, text=button_text, variable=var)
    
    text_area.window_create('end', window=check_btn)
    text_area.insert('end', '\n')




text_area.configure(state=tk.DISABLED)





root.mainloop()